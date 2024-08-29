from flask import Flask, render_template, request, send_file
import os
import pandas as pd
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def process_tumour_volume_data(input_file, output_file, mice_per_group):
    data = pd.read_excel(input_file, sheet_name='Prism Horizontal')
    group_start_indices = list(range(4, data.shape[1], mice_per_group))
    group_dataframes = []
    group_names = []

    for idx, group_start in enumerate(group_start_indices):
        group_end = group_start + mice_per_group
        if group_end > data.shape[1]:
            group_end = data.shape[1]
        group_data = data.iloc[2:, [3] + list(range(group_start, group_end))]
        mouse_numbers_group = data.iloc[1, group_start:group_end].tolist()
        group_name = data.iloc[0, group_start]
        group_names.append(group_name)
        group_column = [group_name] + [''] * (group_data.shape[0])
        group_df = pd.DataFrame({
            'Group': group_column,
            'Date': ['Date'] + pd.to_datetime(group_data.iloc[:, 0], errors='coerce', format='%m/%d/%Y').dt.strftime('%m/%d/%Y').tolist(),
            **{f'Mouse {i+1}': [mouse_numbers_group[i]] + group_data.iloc[:, i + 1].tolist() for i in range(len(mouse_numbers_group))}
        })
        group_df = group_df.reindex(list(range(80)), fill_value='')
        group_dataframes.append(group_df)

    try:
        book = load_workbook(output_file)
    except FileNotFoundError:
        book = Workbook()

    if len(group_dataframes) > 12:
        sheet1 = book.create_sheet(title='Tumour volume Data entry')
        start_row = 5
        header = ["Group", "Date"] + [f"Mouse {i+1}" for i in range(mice_per_group)]

        for idx in range(12):
            group_df = group_dataframes[idx]
            for col_index, header_value in enumerate(header, start=1):
                sheet1.cell(row=start_row, column=col_index).value = header_value

            start_row += 1

            for row_index in range(len(group_df)):
                for col_index, col_name in enumerate(group_df.columns, start=1):
                    value = group_df.iloc[row_index, col_index - 1]
                    sheet1.cell(row=start_row + row_index, column=col_index).value = value

            start_row += 79

        sheet2 = book.create_sheet(title='Tumour volume Data entry - Continued')
        start_row = 5
        group_df = group_dataframes[0]
        for col_index, header_value in enumerate(header, start=1):
            sheet2.cell(row=start_row, column=col_index).value = header_value
        start_row += 1
        for row_index in range(len(group_df)):
            for col_index, col_name in enumerate(group_df.columns, start=1):
                value = group_df.iloc[row_index, col_index - 1]
                sheet2.cell(row=start_row + row_index, column=col_index).value = value
        start_row += 79

        group_position = 2
        for idx in range(12, len(group_dataframes)):
            if group_position == 4: 
                start_row += 80
                group_position += 1
            group_df = group_dataframes[idx]
            for col_index, header_value in enumerate(header, start=1):
                sheet2.cell(row=start_row, column=col_index).value = header_value
            start_row += 1
            for row_index in range(len(group_df)):
                for col_index, col_name in enumerate(group_df.columns, start=1):
                    value = group_df.iloc[row_index, col_index - 1]
                    sheet2.cell(row=start_row + row_index, column=col_index).value = value
            start_row += 79
            group_position += 1
    else:
        sheet = book.create_sheet(title='Tumour volume Data entry')
        start_row = 5
        header = ["Group", "Date"] + [f"Mouse {i+1}" for i in range(mice_per_group)]

        for idx, group_df in enumerate(group_dataframes):
            for col_index, header_value in enumerate(header, start=1):
                sheet.cell(row=start_row, column=col_index).value = header_value

            start_row += 1
            for row_index in range(len(group_df)):
                for col_index, col_name in enumerate(group_df.columns, start=1):
                    value = group_df.iloc[row_index, col_index - 1]
                    sheet.cell(row=start_row + row_index, column=col_index).value = value

            start_row += 79 

    book.save(output_file)

def process_body_weight_data(input_file, output_file, mice_per_group):
    body_weight_data = pd.read_excel(input_file, sheet_name='Regress Tool or Individ Anim')

    group_dataframes = []
    current_group = None
    current_group_data = []

    for index, row in body_weight_data.iterrows():
        if pd.notna(row[2]) and row[2].startswith('Group'):
            if current_group is not None:
                group_df = pd.DataFrame(current_group_data)
                group_df = group_df.reindex(list(range(71)), fill_value='')
                group_dataframes.append((current_group, group_df))

            current_group = row[2]
            current_group_data = []

        if current_group:
            relevant_data = row[[3] + list(range(4, 4 + mice_per_group))].tolist()
            current_group_data.append(relevant_data)

    if current_group_data:
        group_df = pd.DataFrame(current_group_data)
        group_df = group_df.reindex(list(range(71)), fill_value='')
        group_dataframes.append((current_group, group_df))

    try:
        book = load_workbook(output_file)
    except FileNotFoundError:
        book = Workbook()
        sheet = book.active
        sheet.title = 'Body weight Data input'
    
    if 'Body weight Data input' in book.sheetnames:
        sheet = book['Body weight Data input']
    else:
        sheet = book.create_sheet(title='Body weight Data input')

    start_row = 1 

    for group_name, group_df in group_dataframes:
        start_row += 3

        for i in range(mice_per_group):
            sheet.cell(row=start_row + 1, column=3 + i).value = f'Mouse {i + 1}'
        for row_index in range(71):
            sheet.cell(row=start_row + 2 + row_index, column=1).value = '' if row_index != 0 else group_name
            for col_index, value in enumerate(group_df.iloc[row_index], start=2):
                if col_index == 2 and isinstance(value, pd.Timestamp):
                    value = value.strftime('%m/%d/%Y')
                sheet.cell(row=start_row + 2 + row_index, column=col_index).value = value
        
        start_row += 69  

    book.save(output_file)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            return "No file part"
        file = request.files['file']
        if file.filename == '':
            return "No selected file"
        if file:
            input_path = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(input_path)

            output_filename = request.form['output_filename']
            if not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'
            
            output_file = os.path.join(UPLOAD_FOLDER, output_filename)
            mice_per_group = int(request.form['mice_per_group'])
            processing_type = request.form['processing_type']

            if processing_type == 'tumour_volume':
                process_tumour_volume_data(input_path, output_file, mice_per_group)
            elif processing_type == 'body_weight':
                process_body_weight_data(input_path, output_file, mice_per_group)

            return send_file(output_file, as_attachment=True)

    return render_template('index.html')

if __name__ == '__main__':
    from waitress import serve
    serve(app, host='0.0.0.0', port=8080)